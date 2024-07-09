# openpyxl 라이브러리에서 Workbook을 import
from openpyxl import Workbook

# Workbook 객체 생성
wb= Workbook()

# 현재 활성화된 워크시트 선택 후 ws 변수에 할당
ws = wb.active

#시트 제목을 '수강생_정보'로 변경
ws.title = "수강생_정보"

#시트의 A1셀에 '이철수'라는 데이터를 입력
ws['A1'] = '이철수'
ws['A2'] = '순희'; ws['B1'] = '이순신'
# 워크북을 '수강생_리스트.xlsx' 엑셀 파일을 저장합니다
wb.save("수강생_리스트.xlsx")

#워크북 닫기
wb.close()
 