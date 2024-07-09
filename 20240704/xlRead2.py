import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 엑셀 파일 경로
file_path = '월별구매고객리스트.xlsx'
output_file_path = '월별구매고객리스트.xlsx'

# 엑셀 파일 읽기
df = pd.read_excel(file_path)

# 엑셀 파일을 openpyxl로 열기
wb = load_workbook(file_path)
ws = wb.active

# 변경할 열의 이름 (예: 'Column1')
column_to_color = '10월'

# 빨간색 글자 설정
red_font = Font(color="FF0000")

# 헤더의 위치 찾기
header_row = 1
column_index = None
for cell in ws[header_row]:
    if cell.value == column_to_color:
        column_index = cell.column
        break

if column_index is not None:
    # 열의 모든 셀에 적용 (헤더 제외)
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        if isinstance(cell.value, str) or isinstance(cell.value, int) or isinstance(cell.value, float):
            cell.font = red_font
else:
    print(f"Column '{column_to_color}' not found in the Excel file.")

# 수정된 엑셀 파일 저장
wb.save(output_file_path)
print(f"Updated file saved as '{output_file_path}'")