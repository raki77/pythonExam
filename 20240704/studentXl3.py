from openpyxl import Workbook
from operator import itemgetter
from openpyxl.styles import Font
 
wb = Workbook()
ws = wb.active
ws.title = "수강생_정보"

# 시트에 추가할 칼럼의 목록을 리스트형으로 'column'이라는 변수에 지정
column = ['번호', '이름', '과목']

# column리스트 목록을 시트 첫 행에 입력
ws.append(column)
 
# 각 행의 데이터 리스트를 포함하는 리스트를 만들어 row에 저장
row = [[1, '이철수', '수학'], [3, '김미소', '영어'], [2, '최학준', '컴퓨터']]
# row리스트를 첫번째 키를 기준으로 정렬
row = sorted(row, key = itemgetter(0))
# row리스트를 for문으로 반복하여 시트에 입력
for data in row:    
    # if data[2] == '수학':
    #     cell = data[2] 
    #     cell.font = Font(color='FF0000', italic=True, bold=True, size=20)        
    ws.append(data)

# '중간평가', '기말평가' 두 개의 시트 추가
wb.create_sheet('중간평가')
wb.create_sheet('기말평가')

wb.save("수강생_리스트3.xlsx")
wb.close()
