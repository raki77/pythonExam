from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = "수강생_정보"

# 시트에 추가할 칼럼의 목록을 리스트형으로 'column'이라는 변수에 지정
column = ['번호', '이름', '과목']

# column리스트 목록을 시트 첫 행에 입력
ws.append(column)

# 시트에 추가할 데이터를 리스트형으로 'row'라는 변수에 지정
row = [1, '이철수', '수학'] 

# append로 row의 목록을 column 아래 행에 입력
ws.append(row)

# '중간평가', '기말평가' 두 개의 시트 추가
wb.create_sheet('중간평가')
wb.create_sheet('기말평가')

wb.save("수강생_리스트2.xlsx")
wb.close()
