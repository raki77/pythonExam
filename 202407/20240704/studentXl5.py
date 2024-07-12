from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "수강생_정보"

# 열로 입력할 데이터를 리스트로 만들어 'data'에 저장
data = [ '이철수', '김미소', '최학준' ]


# 입력할 열의 인덱스 번호를 열이름으로 변환하여 'column_name'에 저장
column_name = get_column_letter(1)

# 데이터를 입력할 열의 각 셀의 위치를 for문으로 생성한후 'data'의 값을 하나씩 입력
for i, value in enumerate(data):
    ws[f"{column_name}{i+1}"] = value           # 파이썬 f-string

wb.save("수강생_리스트5.xlsx")
wb.close()